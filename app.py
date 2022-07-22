from distutils.log import error
from genericpath import isfile
from webdriver_manager.chrome import ChromeDriverManager
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from tkinter import messagebox
import logging
import os


path = "./FT_a_procesar"
contenido = os.listdir(path)
fichas = []

for ficha in contenido:
    if os.path.isfile(os.path.join(path, ficha)) and ficha.endswith(".xlsm"):
        fichas.append(ficha)
print(fichas)

logging.basicConfig(
    filename="app.txt",
    level=logging.INFO,
    format="%(asctime)s:%(levelname)s:%(message)s",
)

# read/write Excel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import time

""" Download chromedriver """
# chrome://version/
# https://chromedriver.storage.googleapis.com/index.html
# pip install openpyxl


logging.info("...Iniciando...")

# driver_service = Service(executable_path="./selenium-driver/chromedriver.exe")
driver = webdriver.Chrome(ChromeDriverManager().install())
driver.maximize_window()
# URL
driver.get("http://vpn.grisino.com:8001/maertest")


class Login:
    def __init__(self, user, password):
        self.user = user
        self.password = password

    def login(self):
        try:
            logging.info("Iniciando sesion..")
            login_user = WebDriverWait(driver, 10).until(
                expected_conditions.presence_of_element_located(
                    (By.ID, "ext-comp-1002")
                )
            )

            login_user.send_keys(self.user)

            password_user = WebDriverWait(driver, 10).until(
                expected_conditions.presence_of_element_located(
                    (By.ID, "ext-comp-1004")
                )
            )
            password_user.send_keys(self.password)

            login_btn = driver.find_element(By.ID, "ext-gen31")
            login_btn.click()
            logging.info("Entrando...")
        except (Exception) as error_excepction:
            logging.warning("Error: ", error_excepction)
            messagebox.showerror(message=error_excepction, title="Error")


class LoadFile:
    def __init__(self, fichas):
        self.fichas = fichas

    def loop(self, rango_cod_color, lista_cod_color):
        # Iterar por cod de color
        for cod in rango_cod_color:
            for i in cod:
                lista_cod_color.append(i.value)
                print(i.value)

    def loop_cod_color(self, rango_cod_color, lista_cod_color, celda):
        # Iterar por cod de color
        if type(celda).__name__ == "MergedCell":
            print("Combinada, color todos")
            lista_cod_color.append(celda.value)
            return True
        else:
            for cod in rango_cod_color:
                for i in cod:
                    lista_cod_color.append(i.value)
                    print(i.value)
            return False

    def comprobar_y_cargar(
        self,
        actions,
        descripcion_validacion,
        talles,
        lista_cod_color,
        cantidad_insumo_confeccion,
        insumo_confeccion,
    ):
        if "TALLE" in descripcion_validacion:
            logging.info("Cargnado insumo por talle...")
            # Por cada talle cargar...
            for i in talles:
                time.sleep(2)
                if i != None:
                    logging.info("talles disponibles: ", i)
                    print("talles disponibles: ", i)
                    time.sleep(2)
                    for cod_color in lista_cod_color:
                        if cod_color != None:
                            self.load_insumo_por_talle(
                                actions,
                                insumo_confeccion,
                                cod_color,
                                cantidad_insumo_confeccion,
                                i,
                            )
                            print("Carga de insumo por talle finalizada...")
                            logging.info("Carga de insumo por talle finalizada...")
        else:
            for i in lista_cod_color:
                if i != None:
                    self.load_insumo2(
                        actions,
                        insumo_confeccion,
                        i,
                        cantidad_insumo_confeccion,
                    )
                else:
                    pass

    def load_insumo(self, actions, insumo, color_insumo, cantidad):
        if insumo != None:
            logging.info(f"Cargando el insumo {insumo}")
            time.sleep(2)
            actions.send_keys(insumo + "." + color_insumo)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.TAB)
            actions.perform()
            time.sleep(2)
            actions.send_keys(cantidad)
            time.sleep(2)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            time.sleep(2)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
        else:
            actions.send_keys(Keys.ESCAPE)
            actions.perform()
            actions.send_keys(Keys.ESCAPE)
            actions.perform()
            logging.info("Carga de insumo finalizada")

    def load_insumo2(self, actions, insumo, i, cantidad):
        if insumo != None:
            logging.info(f"Cargando el insumo {insumo}")
            time.sleep(2)
            actions.send_keys(insumo + "." + i)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.TAB)
            actions.perform()
            time.sleep(2)
            actions.send_keys(cantidad)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.TAB)
            actions.perform()
        else:
            pass

    def load_insumo_por_talle(self, actions, insumo, color_insumo, cantidad, i):
        if insumo != None and color_insumo != None:
            logging.info(f"Cargando el insumo {insumo}")
            time.sleep(1)
            actions.send_keys(insumo + "." + color_insumo)
            actions.perform()
            time.sleep(1)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            actions.send_keys(Keys.TAB)
            actions.perform()
            time.sleep(2)
            actions.send_keys(cantidad)
            actions.perform()
            time.sleep(3)
            actions.send_keys(Keys.TAB)
            actions.perform()
            time.sleep(3)
            actions.send_keys(color_insumo)
            actions.send_keys(Keys.ENTER)
            actions.perform()
            time.sleep(2)
            print(type(i))
            if i != None:
                if i != "2":
                    actions.send_keys(i)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                else:
                    actions.send_keys(i)
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.TAB)
                    actions.perform()

    def load_new(self):
        try:
            time.sleep(1)
            btn_produccion = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (
                        By.XPATH,
                        "/html/body/div[1]/div[2]/div/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[2]/em/button",
                    )
                )
            )

            btn_produccion.click()

            btn_ficha_tecnica = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (By.LINK_TEXT, "Fichas Técnicas")
                )
            )
            btn_ficha_tecnica.click()

            btn_ficha_tecnica2 = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (By.LINK_TEXT, "Fichas T?cnicas")
                )
            )
            btn_ficha_tecnica2.click()

            btn_maxim_ft = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (
                        By.CLASS_NAME,
                        "x-tool-maximize",
                    )
                )
            )
            btn_maxim_ft.click()
            btn_add_new = WebDriverWait(driver, 35).until(
                expected_conditions.presence_of_element_located(
                    (
                        By.XPATH,
                        "//button[contains(text(),'Agregar')]",
                    )
                )
            )
            time.sleep(2)
            btn_add_new.click()
            logging.info("Nueva ficha tecnica")
            logging.info("reading excel..")

            for self.ficha in self.fichas:
                logging.info(f"Cargando ficha: {self.ficha}")
                wb = load_workbook(f"./FT_a_procesar/{self.ficha}", data_only=True)
                ws = wb.active

                time.sleep(10)

                input_coleccion = driver.find_element(
                    By.XPATH,
                    "/html/body/div[4]/div[2]/div/div/div/div[1]/div[10]/div[2]/div[1]/div/div/div/div/div/div[1]/div[1]/div[2]/div/div[1]/div[2]/div/div[3]/div[1]/div/input",
                )
                actions = ActionChains(driver)
                coleccion = ws["G1"].value
                time.sleep(1)
                input_coleccion.send_keys(coleccion)
                time.sleep(3)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(3)
                actions.send_keys(Keys.TAB)
                actions.perform()
                time.sleep(3)
                producto = ws["B2"].value
                time.sleep(3)
                actions.send_keys(producto)
                time.sleep(2)
                actions.perform()
                actions.send_keys(Keys.ARROW_DOWN)
                actions.perform()
                actions.send_keys(Keys.ENTER)
                actions.perform()
                time.sleep(3)
                actions.send_keys(Keys.TAB)
                actions.perform()
                actions.send_keys(Keys.TAB)
                actions.perform()
                molde = ws["T2"].value
                actions.send_keys(molde)
                actions.perform()
                time.sleep(4)

                btn_add_rule = driver.find_element(
                    By.XPATH,
                    "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[1]/div/table/tbody/tr/td[1]/table/tbody/tr/td[1]/table/tbody/tr[2]/td[2]/em/button",
                )
                time.sleep(5)
                logging.info("Agregando regla - telas corte")
                btn_add_rule.click()
                actions.send_keys("100-CORTE ORIGINAL")
                actions.perform()
                time.sleep(1)
                actions.send_keys(Keys.ENTER)
                actions.perform()
                actions.send_keys(Keys.ESCAPE)
                actions.perform()
                time.sleep(3)

                nueva_entrada = driver.find_element(
                    By.XPATH,
                    "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                )
                time.sleep(5)
                nueva_entrada.click()
                time.sleep(6)

                agregar_insumo = driver.find_element(
                    By.XPATH,
                    "//table[@id='ext-comp-1322']/tbody/tr[2]/td[2]/em/button",
                )
                time.sleep(4)
                agregar_insumo.click()
                logging.info("Agregando insumos telas")
                time.sleep(1)
                actions.send_keys(Keys.TAB)
                actions.perform()

                insumo_1 = ws["I6"].value
                color_inusmo = ws["L7"].value
                color_insumo2 = ws["L9"].value
                cantidad_insumo_1 = str(ws["J6"].value)
                cantidad_insumo_2 = ws["J8"].value

                time.sleep(2)
                self.load_insumo(actions, insumo_1, color_inusmo, cantidad_insumo_1)
                time.sleep(2)
                COLOR1 = ws["L4"].value
                COLOR2 = ws["N4"].value
                insumo_2 = ws["I8"].value
                insumo_3 = ws["I10"].value
                color_insumo3 = ws["L11"].value
                cantidad_insumo_3 = ws["J10"].value
                insumo_4 = ws["I12"].value
                insumo_5 = ws["I14"].value
                insumo_6 = ws["I16"].value
                color_insumo4 = ws["N5"].value
                # XTA004
                color_insumo5 = ws["N5"].value
                # XTD001
                color_insumo6 = ws["N5"].value
                cantidad_insumo_4 = str(ws["J12"].value)
                cantidad_insumo_5 = str(ws["J14"].value)
                cantidad_insumo_6 = str(ws["J16"].value)

                # Si insumo existe.. agregar otro
                # Se puede hacer una fx decoradora -----------------------------------------------------------
                if insumo_2 != None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions, insumo_2, color_insumo2, cantidad_insumo_2
                    )
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_3 != None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions, insumo_3, color_insumo3, cantidad_insumo_3
                    )
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_4 != None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions, insumo_4, color_insumo4, cantidad_insumo_4
                    )

                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_5 != None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions, insumo_5, color_insumo5, cantidad_insumo_5
                    )

                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)

                if insumo_6 != None:
                    agregar_insumo.click()
                    actions.send_keys(Keys.TAB)
                    time.sleep(2)
                    actions.perform()
                    time.sleep(2)
                    self.load_insumo(
                        actions, insumo_6, color_insumo6, cantidad_insumo_6
                    )
                else:
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()
                    logging.info("Carga de insumos terminada")

                time.sleep(2)
                btn_guardar = driver.find_element(
                    By.XPATH,
                    "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[2]/div[1]/table/tbody/tr/td[2]/table/tbody/tr/td[1]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[2]/em/button",
                )
                time.sleep(2)
                btn_guardar.click()
                time.sleep(1)
                btn_si = driver.find_element(
                    By.XPATH, "//button[contains(text(),'Sí')]"
                )
                time.sleep(2)
                btn_si.click()
                btn_ok = driver.find_element(
                    By.XPATH, "//button[contains(text(),'OK')]"
                )
                time.sleep(2)
                btn_ok.click()
                logging.info("Ficha Guardada")
                time.sleep(4)
                btn_close = driver.find_element(
                    By.XPATH,
                    "//div[@id='ext-comp-1473']/div/div/div/div/div",
                )
                time.sleep(3)
                btn_close.click()
                time.sleep(2)
                actions.send_keys(Keys.ESCAPE)
                actions.perform()

                """ ------------------------ ---------------------------------------------------- """

                cod_art_2 = ws["B3"].value

                if cod_art_2 != None:
                    logging.info(f"Cargando segundo codigo de producto: {cod_art_2}")
                    input_coleccion = driver.find_element(
                        By.XPATH,
                        "//*[@id='ext-comp-1252']",
                    )
                    actions = ActionChains(driver)
                    coleccion = ws["G1"].value
                    time.sleep(1)
                    input_coleccion.send_keys(coleccion)
                    time.sleep(3)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(3)
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    time.sleep(3)
                    actions.send_keys(cod_art_2)
                    time.sleep(2)
                    actions.perform()
                    actions.send_keys(Keys.ARROW_DOWN)
                    actions.perform()
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    time.sleep(3)
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    actions.send_keys(Keys.TAB)
                    actions.perform()
                    molde = ws["T2"].value
                    actions.send_keys(molde)
                    actions.perform()

                    time.sleep(5)
                    btn_add_rule.click()
                    time.sleep(1)
                    actions.send_keys("100-CORTE ORIGINAL")
                    actions.perform()
                    time.sleep(1)
                    actions.send_keys(Keys.ENTER)
                    actions.perform()
                    actions.send_keys(Keys.ESCAPE)
                    actions.perform()

                    time.sleep(10)
                    logging.info("Agregando entrada")
                    nueva_entrada2 = driver.find_element(
                        By.XPATH,
                        "/html/body/div[4]/div[2]/div/div/div/div[1]/div[8]/div[2]/div[1]/div/div/div/div/div/div[1]/div[2]/div/div/div/div/div[2]/div/div[1]/div[2]/div[1]/div/table/tbody/tr/td[3]/div/span/table/tbody/tr[2]/td[2]/em/button",
                    )
                    time.sleep(10)
                    nueva_entrada2.click()
                    time.sleep(2)
                    agregar_insumo2 = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1515']/tbody/tr[2]/td[2]/em/button",
                    )
                    # "//button[starts-with(@id,'ext-gen') and @type='button' and contains(text(),'Agregar')]"
                    time.sleep(2)
                    agregar_insumo2.click()
                    logging.info("Cargando insumos...")
                    time.sleep(4)
                    actions.send_keys(Keys.TAB)
                    actions.perform()

                    insumo_1 = ws["I6"].value
                    color_inusmo = ws["L7"].value
                    color_insumo2 = ws["L9"].value
                    cantidad_insumo_1 = str(ws["K6"].value)
                    cantidad_insumo_2 = str(ws["K8"].value)

                    time.sleep(2)
                    self.load_insumo(actions, insumo_1, color_inusmo, cantidad_insumo_1)
                    time.sleep(2)
                    COLOR1 = ws["L4"].value
                    COLOR2 = ws["N4"].value
                    insumo_2 = ws["I8"].value
                    insumo_3 = ws["I10"].value
                    color_insumo3 = ws["L11"].value
                    cantidad_insumo_3 = ws["K10"].value
                    insumo_4 = ws["I12"].value
                    insumo_5 = ws["I14"].value
                    insumo_6 = ws["I16"].value
                    color_insumo4 = ws["N5"].value
                    # XTA004
                    color_insumo5 = ws["N5"].value
                    # XTD001
                    color_insumo6 = ws["N5"].value
                    cantidad_insumo_4 = str(ws["K12"].value)
                    cantidad_insumo_5 = str(ws["K14"].value)
                    cantidad_insumo_6 = str(ws["K16"].value)

                    # Si insumo existe.. agregar otro
                    # Se puede hacer una fx decoradora -----------------------------------------------------------
                    if insumo_2 != None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_2, color_insumo2, cantidad_insumo_2
                        )
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(2)

                    if insumo_3 != None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_3, color_insumo3, cantidad_insumo_3
                        )
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(2)

                    if insumo_4 != None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_4, color_insumo4, cantidad_insumo_4
                        )

                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(2)

                    if insumo_5 != None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_5, color_insumo5, cantidad_insumo_5
                        )

                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(2)

                    if insumo_6 != None:
                        agregar_insumo2.click()
                        actions.send_keys(Keys.TAB)
                        time.sleep(2)
                        actions.perform()
                        time.sleep(2)
                        self.load_insumo(
                            actions, insumo_6, color_insumo6, cantidad_insumo_6
                        )
                    else:
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        actions.send_keys(Keys.ESCAPE)
                        actions.perform()
                        logging.info("Carga de insumos terminada")

                    time.sleep(4)
                    btn_guardar = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1302']/tbody/tr[2]/td[2]/em/button",
                    )
                    time.sleep(2)
                    btn_guardar.click()
                    time.sleep(1)
                    btn_si = driver.find_element(
                        By.XPATH, "//button[contains(text(),'Sí')]"
                    )
                    time.sleep(2)
                    btn_si.click()
                    logging.info("Ficha Guardada")
                    btn_ok2 = driver.find_element(
                        By.XPATH,
                        "//table[@id='ext-comp-1466']/tbody/tr[2]/td[2]/em/button",
                    )
                    time.sleep(2)
                    btn_ok2.click()
                    time.sleep(2)
                    btn_close2 = driver.find_element(
                        By.XPATH, "//div[@id='ext-comp-1637']/div/div/div/div/div"
                    )
                    btn_close2.click()
                    time.sleep(1)
                    messagebox.showinfo(message="Carga Finalizada", title="Info")
                    logging.info("Carga Finalizada")
                else:
                    messagebox.showinfo(message="Carga Finalizada", title="Info")
                    logging.info("Carga Finalizada")

            # -------------------------------------------------- ---------------------------------------------------------------------------
        except (Exception) as error_excepction:
            logging.info("Error: ", error_excepction)
            messagebox.showerror(message=error_excepction, title="Error")
            print(error_excepction)


log = Login("Gfrassetti", "Guido")
log.login()

load = LoadFile(fichas)
load.load_new()
